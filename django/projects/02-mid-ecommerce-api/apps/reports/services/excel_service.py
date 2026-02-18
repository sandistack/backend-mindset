"""
Excel export service using openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class ExcelService:
    """Service for generating Excel files with styling."""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    def set_title(self, title):
        """Set worksheet title."""
        self.ws.title = title[:31]  # Excel has 31 char limit
        return self
    
    def set_headers(self, headers):
        """
        Set header row with styling.
        
        Args:
            headers: List of header strings
        """
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Freeze header row
        self.ws.freeze_panes = 'A2'
        return self
    
    def add_rows(self, data):
        """
        Add data rows with automatic formatting.
        
        Args:
            data: List of lists (rows)
        """
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')
                
                # Format dates
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM'
                # Format currency (columns after 3rd typically contain amounts)
                elif isinstance(value, (int, float)) and col_num > 3:
                    cell.number_format = '#,##0'
        return self
    
    def auto_width(self):
        """Auto-adjust column widths based on content."""
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in self.ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with min 10, max 50
            adjusted_width = min(max(max_length + 2, 10), 50)
            self.ws.column_dimensions[column].width = adjusted_width
        return self
    
    def add_summary_row(self, label, value, row_num=None):
        """
        Add a summary row (e.g., totals).
        
        Args:
            label: Label text
            value: Value to display
            row_num: Row number (auto if None)
        """
        if row_num is None:
            row_num = self.ws.max_row + 2
        
        # Label in first column
        label_cell = self.ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True)
        
        # Value in last column
        value_cell = self.ws.cell(row=row_num, column=self.ws.max_column, value=value)
        value_cell.font = Font(bold=True)
        value_cell.number_format = '#,##0'
        
        return self
    
    def get_buffer(self):
        """Return BytesIO buffer for HTTP response."""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def save(self, filepath):
        """Save to file."""
        self.wb.save(filepath)
        return self
